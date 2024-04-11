import openpyxl
import pandas as pd
import hashlib
import xlsxwriter

# TOKENISATION SMART CONTRACT
def tokeniser(image_file, data1, data2, data3, data4, data5):
    # Create a new DataFrame with the new data
    try:
        workbook  = openpyxl.load_workbook('3PL.xlsx')
        worksheet = workbook.active
    except FileNotFoundError:
        workbook  = openpyxl.Workbook()
        worksheet = workbook.active
    # if workbook.sheetnames:
    # else:
    #     worksheet = workbook.create_sheet(title="Sheet1")
    #     worksheet = workbook.active
        worksheet[f'A1']= 'ID'
        worksheet[f'B1']= 'Data1'
        worksheet[f'C1']= 'Data2'
        worksheet[f'D1']= 'Data3'
        worksheet[f'E1']= 'Data4'
        worksheet[f'F1']= 'Data5'

    next_row = worksheet.max_row + 1

    image_id = hashlib.sha256(image_file.read()).hexdigest() if not hasattr(image_file, 'name') else hashlib.sha256(image_file.name.encode()).hexdigest()
    worksheet[f'A{next_row}'] = image_id
    worksheet[f'B{next_row}'] = data1
    worksheet[f'C{next_row}'] = data2
    worksheet[f'D{next_row}'] = data3
    worksheet[f'E{next_row}'] = data4
    worksheet[f'F{next_row}'] = data5

    workbook.save('3PL.xlsx')